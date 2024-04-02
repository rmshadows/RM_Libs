#!/usr/bin/python3
from openpyxl import Workbook
from openpyxl import load_workbook

"""
文档：https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-one-cell

#### 访问单元格
访问单元格：ws['A4'] 或者 ws.cell(row=4, column=2, value=10)
多个单元格：cell_range = ws['A1':'C2']
>>> colC = ws['C']
>>> col_range = ws['C:D']
>>> row10 = ws[10]
>>> row_range = ws[5:10]
或者：
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
...    for cell in row:
...        print(cell)
和
>>> for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
...     for cell in col:
...         print(cell)
注意：For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.

Because of this feature, scrolling through cells instead of accessing them directly will create them all in memory, even if you don’t assign them a value.
Something like
>>> for x in range(1,101):
...        for y in range(1,101):
...            ws.cell(row=x, column=y)
will create 100x100 cells in memory, for nothing.
---
For performance reasons the Worksheet.columns property is not available in read-only mode.
If you need to iterate through all the rows or columns of a file, you can instead use the Worksheet.rows property:
>>> ws = wb.active
>>> ws['C9'] = 'hello world'
>>> tuple(ws.rows)
((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
(<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
(<Cell Sheet.A3>, <Cell Sheet.B3>, <Cell Sheet.C3>),
(<Cell Sheet.A4>, <Cell Sheet.B4>, <Cell Sheet.C4>),
(<Cell Sheet.A5>, <Cell Sheet.B5>, <Cell Sheet.C5>),
(<Cell Sheet.A6>, <Cell Sheet.B6>, <Cell Sheet.C6>),
(<Cell Sheet.A7>, <Cell Sheet.B7>, <Cell Sheet.C7>),
(<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
(<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))

or the Worksheet.columns property:

>>> tuple(ws.columns)
((<Cell Sheet.A1>,
<Cell Sheet.A2>,
<Cell Sheet.A3>,
<Cell Sheet.A4>,
<Cell Sheet.A5>,
<Cell Sheet.A6>,
...
<Cell Sheet.B7>,
<Cell Sheet.B8>,
<Cell Sheet.B9>),
(<Cell Sheet.C1>,
<Cell Sheet.C2>,
<Cell Sheet.C3>,
<Cell Sheet.C4>,
<Cell Sheet.C5>,
<Cell Sheet.C6>,
<Cell Sheet.C7>,
<Cell Sheet.C8>,
<Cell Sheet.C9>))

If you just want the values from a worksheet you can use the Worksheet.values property. This iterates over all the rows in a worksheet but returns just the cell values:

for row in ws.values:
   for value in row:
     print(value)


"""


def createNewWorkbook(firstSheetName=None):
    """
    新建workbook但不是新建一个实际的文件
    获取第一张自动生成的表wb.active
    Args:
        firstSheetName: 第一Sheet表的名称
    Returns:

    """
    if firstSheetName is None:
        return Workbook()
    else:
        wb = Workbook()
        wb.active.title = firstSheetName
        return wb


def createSheet(wb, sheetName, position=None):
    """
    新增表格
    position = None # insert at the end (default)
     0) # insert at first position
     -1) # insert at the penultimate position
    默认名称：Sheet, Sheet1, Sheet2
    访问：wb["New Title"]
    Args:
        wb:
        sheetName: sheet名
        position: 位置

    Returns:
        返回新增的worksheet
    """
    if position is None:
        return wb.create_sheet(sheetName)
    else:
        return wb.create_sheet(sheetName, position)


def getSheetNames(wb):
    """
    返回Sheets名字
    Args:
        wb:workbook

    Returns:

    """
    return wb.sheetnames


def copyWorksheet(wb, ws):
    """
    复制工作表
    Args:
        wb: 复制到某workbook
        ws: 要被复制的worksheet

    Returns:

    """
    return wb.copy_worksheet(ws)


def readRows(ws, min_row, max_row, min_col, max_col, valueOnly=True):
    """
    读行
    Args:
        ws: worksheet
        min_row: 开始行
        max_row: 结束行
        min_col: 开始列
        max_col: 结束列
        valueOnly: 是否仅返回值
    Returns:
        <Cell Sheet1.A1>
        <Cell Sheet1.A2>
    """
    cs = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=valueOnly):
        for cell in row:
            print(cell)
            cs.append(cell)
    return cs


def readColumns(ws, min_col, max_col, min_row, max_row, valueOnly=True):
    """
        读列
        Args:
            ws: worksheet
            min_col: 开始列
            max_col: 结束列
            min_row: 开始行
            max_row: 结束行
            valueOnly: 是否仅返回值
        Returns:
            <Cell Sheet1.A1>
            <Cell Sheet1.A2>
        """
    cs = []
    for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=valueOnly):
        for cell in col:
            print(cell)
            cs.append(cell)
    return cs


def getCellRange(ws, cell1, cell2):
    """
    返回一个范围内的单元格
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]
    Args:
        ws: worksheet
        ws['A1':'C2']
        cell1: 起始
        cell2: 末尾

    Returns:

    """
    return ws[cell1:cell2]


def getRowOrColumns(ws, name):
    """
    返回行或者列
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]
    Args:
        ws: worksheet
        ws['A1':'C2']
        cell1: 起始
        cell2: 末尾

    Returns:

    """
    return ws[name]

def justReadOneRow(ws, row, valueOnly=True):
    """
    读1行,50列
    Args:
        ws: worksheet
        row: 行
        valueOnly: 是否仅返回值
    Returns:
        <Cell Sheet1.A1>
        <Cell Sheet1.A2>
    """
    cs = []
    for row in ws.iter_rows(min_row=row, max_row=row, min_col=0, max_col=50, values_only=valueOnly):
        for cell in row:
            print(cell)
            cs.append(cell)
    return cs


def justReadOneColumn(ws, min_col, max_col, min_row, max_row, valueOnly=True):
    """
        读列
        Args:
            ws: worksheet
            min_col: 开始列
            max_col: 结束列
            min_row: 开始行
            max_row: 结束行
            valueOnly: 是否仅返回值
        Returns:
            <Cell Sheet1.A1>
            <Cell Sheet1.A2>
        """
    cs = []
    for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=valueOnly):
        for cell in col:
            print(cell)
            cs.append(cell)
    return cs


def loadExcel(filename):
    """
    加载现有的Excel表格
    Args:
        filename:Excel文件名

    Returns:
        wb
    """
    return load_workbook(filename=filename)


if __name__ == '__main__':
    wb = loadExcel("1.xlsx")
    ws = wb[getSheetNames(wb)[0]]
    readColumns(ws, 2, 4, 3, 5, False)
    print("Hello World")
